from __future__ import annotations

from io import BytesIO

import folium
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from shapely.geometry import MultiPolygon, Polygon


class GeohashMapRenderer:
    def render(self, geometry, geohashes: list[str]) -> str:
        center = geometry.representative_point()
        map_obj = folium.Map(
            location=[center.y, center.x],
            zoom_start=11,
            tiles="CartoDB positron",
        )

        for geohash in geohashes:
            self._add_geohash_rectangle(map_obj, geohash)

        self._add_geometry_outline(map_obj, geometry)
        return map_obj.get_root().render()

    def _add_geohash_rectangle(self, map_obj: folium.Map, geohash: str) -> None:
        min_lon, min_lat, max_lon, max_lat = self._decode_bbox(geohash)
        bounds = [(min_lat, min_lon), (max_lat, max_lon)]
        color = "#1f77b4" if len(geohash) == 4 else "#ff7f0e"

        folium.Rectangle(
            bounds=bounds,
            color=color,
            weight=1,
            fill=True,
            fill_opacity=0.18,
            tooltip=geohash,
        ).add_to(map_obj)

    def _add_geometry_outline(self, map_obj: folium.Map, geometry) -> None:
        if isinstance(geometry, Polygon):
            coords = [(lat, lon) for lon, lat in geometry.exterior.coords]
            folium.PolyLine(
                coords,
                color="black",
                weight=3,
                opacity=0.9,
                tooltip="Contorno",
            ).add_to(map_obj)
            return

        if isinstance(geometry, MultiPolygon):
            for polygon in geometry.geoms:
                self._add_geometry_outline(map_obj, polygon)

    def _decode_bbox(self, geohash: str) -> tuple[float, float, float, float]:
        import pygeohash as pgh

        lat, lon, lat_err, lon_err = pgh.decode_exactly(geohash)
        return (lon - lon_err, lat - lat_err, lon + lon_err, lat + lat_err)


class GeohashSpreadsheetRenderer:
    def render(self, geohashes: list[str]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "geohashes"
        worksheet["A1"] = "geohash"

        for index, geohash in enumerate(geohashes, start=2):
            worksheet[f"A{index}"] = geohash

        worksheet.column_dimensions[get_column_letter(1)].width = 12

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
